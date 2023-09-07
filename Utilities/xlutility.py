import openpyxl
from openpyxl.styles import PatternFill
def getRow_count(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return (sheet.max_row)
def getColumn_count(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return (sheet.max_coloum)

def ReadData(file,SheetName,RowNum,ColNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(RowNum,ColNum).value
def WriteData(file,SheetName,RowNum,ColNum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    sheet.cell(RowNum,ColNum).value=data
    workbook.save(file)

def FillGreenColor(file,SheetName,RowNum,ColNUm):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(RowNum,ColNUm).fill=greenfill
    workbook.save(file)

def FillRedColur(file,SheetName,RowNum,ColNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    redfill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(RowNum,ColNum).fill=redfill
    workbook.save(file)